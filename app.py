import time
from threading import Timer
from PyQt5.QtGui import QFont, QIntValidator, QRegExpValidator, QPixmap
import smtplib, ssl, random, json
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QPushButton,
    QLabel,
    QVBoxLayout,
    QHBoxLayout,
    QWidget,
    QGroupBox,
    QFormLayout,
    QLineEdit,
    QDialogButtonBox,
    QCalendarWidget,
    QComboBox, QInputDialog, QMessageBox, QGridLayout, QScrollArea)
from PyQt5.QtCore import Qt, QRegExp, QDate, QRect, QTimer
import datetime
s = None
from PyQt5 import QtWidgets, QtCore, QtGui
import sys
import openpyxl as xl

teachers = xl.load_workbook("teachers.xlsx")
teachers = teachers["Sheet1"]
teachers = eval(teachers["A1"].value)

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setGeometry(QtCore.QRect(120, 80, 531, 451))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(16)
        self.groupBox.setFont(font)
        self.groupBox.setObjectName("groupBox")
        self.label = QtWidgets.QLabel(self.groupBox)
        self.label.setGeometry(QtCore.QRect(10, 30, 231, 181))
        self.label.setText("")
        self.label.setPixmap(QtGui.QPixmap("image2.jpg"))
        self.label.setObjectName("label")
        self.pushButton = QtWidgets.QPushButton(self.groupBox)
        self.pushButton.setGeometry(QtCore.QRect(10, 212, 231, 31))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet("background-color: rgb(0, 255, 0);")
        self.pushButton.setObjectName("pushButton")
        self.label_2 = QtWidgets.QLabel(self.groupBox)
        self.label_2.setGeometry(QtCore.QRect(270, 30, 221, 181))
        self.label_2.setText("")
        self.label_2.setPixmap(QtGui.QPixmap("image3.jpg"))
        self.label_2.setObjectName("label_2")
        self.pushButton_2 = QtWidgets.QPushButton(self.groupBox)
        self.pushButton_2.setGeometry(QtCore.QRect(270, 212, 211, 31))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setStyleSheet("background-color: rgb(255, 0, 0);")
        self.pushButton_2.setObjectName("pushButton_2")
        self.label_3 = QtWidgets.QLabel(self.groupBox)
        self.label_3.setGeometry(QtCore.QRect(10, 250, 231, 161))
        self.label_3.setText("")
        self.label_3.setPixmap(QtGui.QPixmap("image4.jpg"))
        self.label_3.setObjectName("label_3")
        self.pushButton_3 = QtWidgets.QPushButton(self.groupBox)
        self.pushButton_3.setGeometry(QtCore.QRect(10, 412, 201, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setStyleSheet("background-color: rgb(0, 0, 255);")
        self.pushButton_3.setObjectName("pushButton_3")
        self.label_4 = QtWidgets.QLabel(self.groupBox)
        self.label_4.setGeometry(QtCore.QRect(270, 250, 221, 161))
        self.label_4.setText("")
        self.label_4.setPixmap(QtGui.QPixmap("image5.jpg"))
        self.label_4.setObjectName("label_4")
        self.pushButton_4 = QtWidgets.QPushButton(self.groupBox)
        self.pushButton_4.setGeometry(QtCore.QRect(270, 410, 221, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.pushButton_4.setFont(font)
        self.pushButton_4.setStyleSheet("background-color: rgb(255, 255, 0);")
        self.pushButton_4.setObjectName("pushButton_4")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 777, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.groupBox.setTitle(_translate("MainWindow", "Welcome"))
        self.pushButton.setText(_translate("MainWindow", "View Curriculum"))
        self.pushButton_2.setText(_translate("MainWindow", "Register a child"))
        self.pushButton_3.setText(_translate("MainWindow", "View database of your children"))
        self.pushButton_4.setText(_translate("MainWindow", "View Assignments(Coming Soon)"))

class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        self.s = None
        self.scroll = QScrollArea()
        self.setStyleSheet("background-image:url(ona.jpg)")
        self.wb = xl.load_workbook("Country-codes.xlsx")
        self.codes = {}

        for i in range(1, 247):
            self.codes[str(self.wb["www.pier2pier.com"]["A"+str(i)].value)] = str(self.wb["www.pier2pier.com"]["B"+str(i)].value)
        self.widget = QWidget()
        self.question = QLabel(self.widget)
        self.bt1 = QPushButton("", self.widget)
        self.bt2 = QPushButton("", self.widget)
        self.bt3 = QPushButton("", self.widget)
        self.bt4 = QPushButton("", self.widget)
        self.bt5 = QPushButton("", self.widget)
        self.nexter = QPushButton("Next", self.widget)
        self.nexter.hide()
        # self.wid2 = QWidget()
        #dict
        self.children = {}
        # self.vbox = QVBoxLayout()
        # self.hbox = QHBoxLayout()
        # self.wid2.setLayout(self.hbox)
        # self.vbox.setAlignment(Qt.AlignCenter)
        #self.hbox.setAlignment(Qt.AlignHCenter)
        # self.widget.setLayout(self.vbox)
        self.setCentralWidget(self.widget)
        self.setWindowTitle("GenChamps")
        self.setGeometry(100, 100, 1280, 616)
        #self.setStyleSheet("background-image: url(tree.jpg)")
        self.w = None
        self.points = 0
        # self.lkk = QLabel("Are you a child or teacher:", self)
        # self.lkk.setStyleSheet("background-color: none;")
        font = QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(24)
        # self.lkk.setFont(font)
        # self.lkk.setAlignment(Qt.AlignBottom)
        self.child = QPushButton("Child", self.widget)
        self.child.setStyleSheet("background-color: rgb(16, 255, 32);")
        self.child.setGeometry(QRect(380, 240, 141, 71))
        self.child.setFont(font)
        self.teacher = QPushButton("Teacher", self.widget)
        self.teacher.setStyleSheet("background-color: rgb(32, 80, 255);")
        self.teacher.setGeometry(QRect(594, 242, 131, 71))
        self.teacher.setFont(font)
        # self.hbox.addWidget(self.child)
        # self.hbox.addWidget(self.teacher)
        # self.vbox.addWidget(self.lkk)
        # self.vbox.addWidget(self.wid2)
        self.child.clicked.connect(self.startMenu)
        self.teacher.clicked.connect(self.teacher_page)

        self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(self.widget)

        self.setCentralWidget(self.scroll)
    def setup(self):
        self.c = random.randint(100000, 999999)
        name, done1 = QInputDialog.getText(
            self, 'Input Dialog', 'Enter your email:')
        true = False
        for i in teachers.keys():
            if teachers[str(i)]["email"] == str(name):
                true = True
                break
            else:
                continue
        try:
            if true == True:
                port = 465  # For SSL
                smtp_server = "smtp.gmail.com"
                sender_email = "orieozichi@gmail.com"  # Enter your address
                receiver_email = str(name)  # Enter receiver address
                password = "growingbetter985"
                message = f"""\
                Subject: Verification\n
    
                You clicked the option for Forgot Password.
                Verification code: {self.c}
                Type this verification code in the GenChamps app"""

                context = ssl.create_default_context()
                with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
                    server.login(sender_email, password)
                    server.sendmail(sender_email, receiver_email, message)
                code, good = QInputDialog.getInt(
                    self, "Info", "Please check your email and you will see a verification number\n\nType the verification number here:"
                )
                if good:
                    your_code = int(code)
                    if your_code != self.c:
                        QMessageBox.about(self, "Error", "You put in a wrong number")
                    else:
                        self.welcometeacher()
            else:
                QMessageBox.about(self, "Error", "Your email is not in our database")
        except:
            QMessageBox.about(self, "Error", "There is an error\nIt is either you entered an incorrect email or that email does not exist")
    def welcometeacher(self):
        # ui = Ui_MainWindow()
        # if self.s is None:
        #     self.s = QtWidgets.QMainWindow()
        #     ui.setupUi(self.s)
        #     self.s.show()
        # else:
        #     self.s.close()
        #     self.s = None
        # s.show()
        #name
        #dateofbirth
        #phonene
        #email
        if True:
            self.todays = QDate.currentDate()
            self.todays = [self.todays.month(), self.todays.day()]
            self.nameofteacher = self.data["name"]
            self.students = self.data["students"]
            self.formgroup.hide()
            self.labelol = QGroupBox(f"Welcome {self.nameofteacher}", self.widget)
            self.labelol.show()
            self.fonts = QFont("Times", 20)
            self.fonts.setItalic(True)
            self.labelol.setFont(self.fonts)
            self.labelol.setGeometry(QRect(400, 80, 531, 451))
            self.labelt = QtWidgets.QLabel(self.labelol)
            self.labelt.setGeometry(QtCore.QRect(10, 30, 231, 181))
            self.labelt.setText("")
            self.labelt.show()
            self.labelt.setPixmap(QtGui.QPixmap("image2.jpg"))
            self.pushButton = QtWidgets.QPushButton(self.labelol)
            self.pushButton.setGeometry(QtCore.QRect(10, 212, 231, 31))
            self.pushButton.setText("View Curriculum")
            self.pushButton.show()
            self.pushButton.clicked.connect(self.view)
            font = QtGui.QFont()
            font.setFamily("Times New Roman")
            font.setPointSize(14)
            self.pushButton.setFont(font)
            self.pushButton.setStyleSheet("background-color: rgb(0, 255, 0);")
            self.label_2 = QtWidgets.QLabel(self.labelol)
            self.label_2.setGeometry(QtCore.QRect(270, 30, 221, 181))
            self.label_2.setText("")
            self.label_2.show()
            self.label_2.setPixmap(QtGui.QPixmap("image3.jpg"))
            self.pushButton_2 = QtWidgets.QPushButton(self.labelol)
            self.pushButton_2.setGeometry(QtCore.QRect(270, 212, 211, 31))
            font = QtGui.QFont()
            self.pushButton_2.show()
            font.setFamily("Times New Roman")
            font.setPointSize(12)
            self.pushButton_2.setFont(font)
            self.pushButton_2.setText("Register a child")
            self.pushButton_2.setStyleSheet("background-color: rgb(255, 0, 0);")
            self.pushButton_2.clicked.connect(self.register_child)
            self.label_3 = QtWidgets.QLabel(self.labelol)
            self.label_3.setGeometry(QtCore.QRect(10, 250, 231, 161))
            self.label_3.setText("")
            self.label_3.show()
            self.label_3.setPixmap(QtGui.QPixmap("image4.jpg"))
            self.label_3.show()
            self.pushButton_3 = QtWidgets.QPushButton(self.labelol)
            self.pushButton_3.setGeometry(QtCore.QRect(10, 412, 201, 31))
            font = QtGui.QFont()
            self.pushButton_3.show()
            self.pushButton_3.clicked.connect(self.see_database)
            self.pushButton_3.setText("View database of your children")
            font.setPointSize(12)
            self.pushButton_3.setFont(font)
            self.pushButton_3.setStyleSheet("background-color: rgb(0, 0, 255);")
            self.label_4 = QtWidgets.QLabel(self.labelol)
            self.label_4.setGeometry(QtCore.QRect(270, 250, 221, 161))
            self.label_4.setText("")
            self.label_4.show()
            self.label_4.setPixmap(QtGui.QPixmap("image5.jpg"))
            self.pushButton_4 = QtWidgets.QPushButton(self.labelol)
            self.pushButton_4.setGeometry(QtCore.QRect(270, 410, 221, 31))
            font = QtGui.QFont()
            font.setPointSize(12)
            self.pushButton_4.setFont(font)
            self.pushButton_4.setText("View Assignments(Coming Soon)")
            self.pushButton_4.setStyleSheet("background-color: rgb(255, 255, 0);")
            self.pushButton_4.show()
            self.pushButton_4.clicked.connect(self.comers)
        #     self.bt2 = QPushButton("View Curriculum", self.widget)
        #     self.bt2.clicked.connect(self.view)
        #     self.bt2.setFont(QFont("Times New Roman", 12))
        #     self.bt2.setGeometry(10, 200, 231, 23)
        #     self.bt2.setStyleSheet("background-color: blue;")
        #     self.bt2.show()
        #
        #     self.lake = QLabel(self.widget)
        #     self.lake.setGeometry(10, 20, 231, 171)
        #     self.lake.setPixmap(QPixmap("image2.jpg"))
        #     self.lake.show()
        #
        #     self.bt1 = QPushButton("Register a child", self)
        #     self.bt1.clicked.connect(self.register_child)
        #     self.bt1.setFont(QFont("Times", 12))
        #     self.bt1.setStyleSheet("background-color: red;")
        #
        #     self.bt3 = QPushButton("View the database of your children", self)
        #     self.bt3.clicked.connect(self.see_database)
        #     self.bt3.setFont(QFont("Times", 12))
        #     self.bt3.setFixedSize(250, 60)
        #     self.bt3.setStyleSheet("background-color: green;")
        #
        #     self.bt4 = QPushButton("View Assignments\n  Coming Soon", self)
        #     self.bt4.clicked.connect(self.comers)
        #     self.bt4.setFont(QFont("Times", 12))
        #     self.bt4.setFixedSize(250, 60)
        #     self.bt4.setStyleSheet("background-color: orange;")

            # self.gdpt.addWidget(self.bt1, 0,0)
            # self.gdpt.addWidget(self.bt2, 0,1)
            # self.gdpt.addWidget(self.bt3, 1,0)
            # self.gdpt.addWidget(self.bt4, 1,1)
            #
            # self.labelol.setLayout(self.gdpt)

    def exitpanacon(self):
        self.labelol.hide()

    def comers(self):
        QMessageBox.about(self, "Coming Soon", "This option is going to appear sometime later")

    def see_database(self):
        self.labelol.hide()
        """"
        Ifechi:
            Name: Ifechi
            Date of birth: 7/12/2021
            Parent's Email: bashade.orie@gmail.com
            Phone: ...
            Class: ...
        """
        self.messages = "Child Data"
        self.information = ""
        students = self.data["students"]
        self.kfc = xl.load_workbook("children.xlsx")
        self.sheets = self.kfc["Sheet1"]
        self.needed_cell = eval(self.sheets["A1"].value)
        for i in students:
            child_data = self.needed_cell[i]
            self.information += f"{child_data['name']}:\n  Name: {child_data['name']}\n  Date of birth: {child_data['date of birth']}\n  Parent's Email: {child_data['email']}\n  Parent's Phone No.: {child_data['phone']}\n  Class: {child_data['classe']}\n"
        self.result = QLabel(self.information, self.widget)
        self.title = QLabel(self.messages, self.widget)
        self.title.setFixedHeight(40)
        self.result.setGeometry(0, self.title.y()+100, self.information.count('\n') * 30^2, self.information.count('\n')*30^2)
        self.exitbt = QPushButton("Exit", self.widget)
        self.exitbt.setStyleSheet("background-color: black; color: white;")
        self.exitbt.setGeometry(0, self.result.height()+700, 131, 40)

        self.result.setFont(QFont("Times", 30))
        self.title.setFont(QFont("Times", 30))
        self.exitbt.setFont(QFont("Times", 30))

        self.exitbt.clicked.connect(self.goaway)
        bar = self.scroll.verticalScrollBar()
        bar.rangeChanged.connect(lambda: bar.setValue(bar.maximum()))
        self.title.show()
        self.result.show()
        self.exitbt.show()
    def goaway(self):
        self.result.hide()
        self.title.hide()
        self.exitbt.hide()
        self.labelol.show()

    def view(self):
        self.labelol.hide()
        self.span = 0
        self.labelk = QLabel(self.widget)
        self.labelk.setPixmap(QPixmap(f"July/0001.jpg"))
        self.labelk.setGeometry(100, 0, 414, 500)
        self.labelk.showNormal()
        self.exitbht = QPushButton("Exit", self.widget)
        self.exitbht.setStyleSheet("background-color: black; color: white;")
        self.exitbht.clicked.connect(self.fgct)
        self.exitbht.setGeometry(100, self.labelk.height()+100, 131, 40)
        self.exitbht.show()
    def fgct(self):
        self.labelk.hide()
        self.exitbht.hide()
        self.labelol.show()
    def confirm(self):
        try:
            self.data = teachers[str(self.username.text())]
            password = str(self.password.text())
            if self.data["password"] != password:
                QMessageBox.about(self, "Error", "Wrong password. Please try again.")
                self.true2 = False
            else:
                self.welcometeacher()
                self.true2 = True
        except Exception as exc:
            print(exc)
            QMessageBox.about(self, "Error", "That account does not exist.")
    def register_child(self):
        self.labelol.hide()
        self.new_box = QGroupBox("Register a child", self.widget)
        self.new_box.showNormal()
        self.font2 = QFont("Times", 30)
        self.font2.setItalic(True)
        self.new_box.setFont(self.font2)
        self.new_box.setGeometry(400, 200, 900, 600)
        self.form = QFormLayout()
        self.name = QLineEdit()
        self.phonoo = QLineEdit()
        self.email3 = QLineEdit()
        self.gender = QLineEdit()
        my_regex4 = QRegExp("^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$")
        my_validator4 = QRegExpValidator(my_regex4, self.email3)
        self.email3.setValidator(my_validator4)
        ranges = QIntValidator()
        self.phonoo.setValidator(ranges)
        self.vbox2 = QVBoxLayout()
        self.submiter = QPushButton("Submit", self.widget)
        self.backer = QPushButton("Exit", self.widget)
        self.backer.setFont(QFont("Times", 20))
        self.backer.setStyleSheet("background-color: black; color: white;")
        self.submiter.setFont(QFont("Times", 20))
        self.submiter.setStyleSheet("background-color: green;")
        self.calendar = QCalendarWidget(self.widget)
        self.calendar.setFixedSize(500, 300)
        self.calendar.setFont(QFont('Times', 6))
        date = QDate.currentDate()
        y = date.year()
        m = date.month()
        d = date.day()

        self.calendar.setMaximumDate(date)
        self.vbox2.addWidget(self.calendar)
        self.form.addRow("Child's Name:", self.name)
        self.form.addRow("Child's Gender(f/m):", self.gender)
        self.form.addRow("Child's Date of birth:", self.vbox2)
        self.form.addRow("Parent's Phone Number", self.phonoo)
        self.form.addRow("Parent's Email:", self.email3)
        self.form.addRow("", self.submiter)
        self.new_box.setLayout(self.form)
        self.submiter.clicked.connect(self.gog)
        self.backer.clicked.connect(self.connect)
        try:
            self.submiter.show()
        except:
            pass

    def connect(self):
        self.new_box.hide()
        self.submiter.hide()
        self.backer.hide()
        self.labelol.show()
    def gog(self):
        try:
            first = random.choice(["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z"])
            second = random.choice(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"])
            third = random.choice(["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"])
            fourth = random.choice(["$", "#", "%", "^", "*", "A", "z", "V", "c"])
            fivth = random.choice(["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p", "q", "r", "s", "t","u", "v", "w", "x", "y", "z"])
            sixth = random.choice(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T","U", "V", "W", "X", "Y", "Z"])
            seventh = random.choice(["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"])
            eight = random.choice(["$", "#", "%", "^", "*", "A", "z", "V", "c"])
            self.passcode = first+second+third+fourth+fivth+sixth+seventh+eight
            port = 465  # For SSL


            self.wb1 = xl.load_workbook("teachers.xlsx")
            self.sheet1 = self.wb1["Sheet1"]
            self.teachers = eval(self.sheet1["A1"].value)
            self.teachers[self.nameofteacher]["students"].append(str(self.name.text()))
            self.sheet1["A1"] = str(self.teachers)
            self.wb1.save("C:\\Users\\HP\\PycharmProjects\\untitled\\GenChamps\\teachers.xlsx")
            age = self.calculate_action()
            if age >= 3 and age < 6:
                self.classe = "Preschool"
            elif age >=6 and age < 10:
                self.classe = "Elementary"
            elif age >=10:
                self.classe = "Preeteens"
            elif age<3:
                QMessageBox.about(self, "Error", "The child that is being registered is to be between the age 3 and above")
                pass
            try:
                smtp_server = "smtp.gmail.com"
                sender_email = "orieozichi@gmail.com"  # Enter your address
                receiver_email = str(self.email3.text())  # Enter receiver address
                password = "growingbetter985"
                message = f"""\
                                            Subject: Loveworld Children Church Registration\n
    
                                            Your child,{self.name.text()} has been registered into a class by {self.nameofteacher} into {self.classe} at the Loveworld Curriculum App.
                                            Password: {self.passcode}
                                            If you have not installed the GenChamps App, Get the app on Google Play and on Amazon Appstore.
                                            Then let your child start understanding the word of God."""

                context = ssl.create_default_context()
                with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
                    server.login(sender_email, password)
                    server.sendmail(sender_email, receiver_email, message)
                QMessageBox.about(self, "Congratulations", f"You have registered a {self.classe} child into the GenChamps app")
                self.wb = xl.load_workbook("children.xlsx")
                self.sheet = self.wb["Sheet1"]
                self.children = eval(self.sheet["A1"].value)
                self.children[self.name.text()] = {"name": self.name.text(),
                                                   "date of birth": f"{self.calendar.selectedDate().month()}/{self.calendar.selectedDate().day()}/{self.calendar.selectedDate().year()}",
                                                   "email": self.email3.text(), "phone": self.phonoo.text(),
                                                   "password": str(self.passcode), "country": "", "classe": self.classe}
                self.sheet["A1"] = str(self.children)
                self.connect()
                self.wb.save("C:\\Users\\HP\\PycharmProjects\\untitled\\GenChamps\\children.xlsx")
            except:
                QMessageBox.about(self, "Error", "You need internet connection")
        except Exception as e:
            print(e)
            age = self.calculate_action()
            # if age >= 3 and age < 6:
            #     self.classe = "Preschool"
            #     QMessageBox.about(self, "Class", str(self.classe)+"\n"+self.passcode)
            # elif age >=6 and age < 10:
            #     self.classe = "Elementary"
            #     QMessageBox.about(self, "Class", str(self.classe)+"\n"+self.passcode)
            # elif age >=10:
            #     self.classe = "Preeteens"
            #     QMessageBox.about(self, "Class", str(self.classe)+"\n"+self.passcode)
            # else:
            #     QMessageBox.about(self, "Error", "The child that is being registered is to be between the age 3-12")
            QMessageBox.about(self, "Error", "There is an error in what you have written. It is either:\n->The email that you entered was invalid.\n->You do not have an internet connection.\n->There is a server problem.")
    def checker(self, widgeter):
        if widgeter.text() != self.corr:
            widgeter.setStyleSheet("background-color: red; border: 9px solid black;")
            if self.bt1.text() == self.corr:
                self.bt1.setStyleSheet("background-color: green; border: 9px solid black;")
            elif self.bt2.text() == self.corr:
                self.bt2.setStyleSheet("background-color: green; border: 9px solid black;")
            elif self.bt3.text() == self.corr:
                self.bt3.setStyleSheet("background-color: green; border: 9px solid black;")
            elif self.bt4.text() == self.corr:
                self.bt4.setStyleSheet("background-color: green; border: 9px solid black;")
            else:
                self.bt5.setStyleSheet("background-color: green; border: 9px solid black;")
            if self.stop != True:
                self.nexter.setEnabled(True)
                self.nexter.clicked.connect(self.nextej)
                #self.vbox.addWidget(self.nexter)
            else:
                self.calculate()
        else:
            self.points += 1
            widgeter.setStyleSheet("background-color: green; border: 9px solid black;")
            if self.stop != True:
                self.nexter.setEnabled(True)
                self.nexter.clicked.connect(self.nextej)
                # self.vbox.addWidget(self.nexter)
            else:
                self.calculate()

    def calculate(self):
        if len(self.options) == 5:
            self.question.hide()
            self.bt1.hide()
            self.bt2.hide()
            self.bt3.hide()
            self.bt4.hide()
            self.bt5.hide()
            self.backed = QPushButton("Exit", self.widget)
            self.backed.show()
            self.backed.setGeometry(0, 140, 100, 60)
            self.backed.setFont(QFont("Times", 35))
            self.backed.clicked.connect(self.cancel)
            self.labell = QLabel(f"Congratulations You Have\n{self.points} points", self.widget)
            self.labell.setFont(QFont("Times", 35))
            self.labell.show()


    def cancel(self):
        self.labell.hide()
        self.backed.hide()

        self.label.showNormal()
        self.movebt.showNormal()

    def nextej(self, widgeter):
        if self.next_choice == "*":
            self.addQuestion("Who died for our sins", 1, ["A lamb", "Jesus", "Angels", "Moses", "Elijah"], "Hebrews 9:28", False, "y")
        if self.next_choice == "-":
            self.addQuestion("Where did Jonah go to when God told him to go Ninevah", 3, ["America", "Heaven", "Jerusalem", "Tarshish", "Antioch"], "Jonah 1:3", True, "P")
    def exit(self):
        self.formgroup.hide()
        self.box.hide()
        self.lkk.show()
        self.child.show()
        self.teacher.show()

    def play(self):
        self.next_choice = "*"

        # self.vbox.addWidget(self.question)
        # self.vbox.addSpacing(100)
        # self.vbox.addWidget(self.bt1)
        # self.vbox.addSpacing(40)
        # self.vbox.addWidget(self.bt2)
        # self.vbox.addSpacing(40)
        # self.vbox.addWidget(self.bt3)
        # self.vbox.addSpacing(40)
        # self.vbox.addWidget(self.bt4)
        # self.vbox.addSpacing(40)
        # self.vbox.addWidget(self.bt5)
        # self.vbox.addWidget(self.nexter)

        try:
            self.label.hide()
            self.movebt.hide()
        except:
            pass
        self.points = 0
        self.addQuestion(
            "Who is called the father of faith.",0, ["Abraham", "Jonah", "Israel", "Jacob", "Samuel"], "Romans 4:12", False, "n"
            )
        self.nextq = ["Who died for our sins",0, ["Abraham", "Jonah", "Israel", "Jacob", "Samuel"], "1 John 5:4", False]
    def addQuestion(self, quest, corr, options, hint, stop, a):
        if a == "y":
            try:
                self.next_choice = "-"
                self.nexter.setEnabled(False)
            except:
                pass
        if a == "z":
            try:
                self.next_choice = "-x"
                self.nexter.setEnabled(False)
            except:
                pass

        self.options = options
        self.stop = stop
        self.aware = True
        self.corr = options[corr]
        self.question.setText(quest+"\n\nHint: "+hint)
        self.question.setStyleSheet("background-color: green; border: 9px solid black;")
        self.question.setGeometry(400, 0, 600, 400)
        self.question.showNormal()
        self.fonte = QFont("Times", 30)
        self.question.setFont(self.fonte)
        if len(options) == 5:
            self.bt1.setText(options[0])
            self.bt2.setText(options[1])
            self.bt3.setText(options[2])
            self.bt4.setText(options[3])
            self.bt5.setText(options[4])

            self.bt1.setGeometry(400, 400, 150, 60)
            self.bt2.setGeometry(400, 450, 150, 60)
            self.bt3.setGeometry(400, 500, 150, 60)
            self.bt4.setGeometry(400, 550, 150, 60)
            self.bt5.setGeometry(400, 600, 150, 60)
            self.nexter.setGeometry(400, 800, 100, 60)
            self.nexter.showNormal()
            self.nexter.setDisabled(True)
            self.bt1.showNormal()
            self.bt2.showNormal()
            self.bt3.showNormal()
            self.bt4.showNormal()
            self.bt5.showNormal()

            self.bt1.clicked.connect(lambda: self.checker(self.bt1))
            self.bt2.clicked.connect(lambda: self.checker(self.bt2))
            self.bt3.clicked.connect(lambda: self.checker(self.bt3))
            self.bt4.clicked.connect(lambda: self.checker(self.bt4))
            self.bt5.clicked.connect(lambda: self.checker(self.bt5))


            self.bt1.setStyleSheet("background-color: blue; border: 9px solid black;")
            self.bt2.setStyleSheet("background-color: blue; border: 9px solid black;")
            self.bt3.setStyleSheet("background-color: blue; border: 9px solid black;")
            self.bt4.setStyleSheet("background-color: blue; border: 9px solid black;")
            self.bt5.setStyleSheet("background-color: blue; border: 9px solid black;")

            # self.vbox.addWidget(self.question)
            # self.vbox.addWidget(self.bt1)
            # self.vbox.addWidget(self.bt2)
            # self.vbox.addWidget(self.bt3)
            # self.vbox.addWidget(self.bt4)
            # self.vbox.addWidget(self.bt5)




    def welcomnewMember(self):
        try:
            self.formgroup.hide()
            self.box.hide()
        except:
            pass
        try:
            self.laying.hide()
            self.submits.hide()
            self.exits.hide()
        except:
            pass
        self.label = QLabel(f"Welcome {self.username.text()}, to the Loveworld Curriclum App.", self)
        self.label.show()
        self.label.setGeometry(self.x()/2, self.y()/2, 1000, 60)
        self.movebt = QPushButton("Play a quiz", self)
        self.movebt.setFixedHeight(80)
        self.movebt.setStyleSheet("background-color: green;")
        self.movebt.clicked.connect(self.play)
        self.movebt.setGeometry(0, 80, 100, 60)
        self.movebt.setFont(QFont("Times New Roman", 14))
        self.movebt.show()
        fonter = QFont("Times", 30)
        fonter.setBold(True)
        self.label.setFont(fonter)
    def teacher_page(self):
        self.setStyleSheet("")
        self.exit1()
        self.formgroup = QGroupBox("Login", self.widget)
        self.formgroup.setGeometry(QRect(400, 120, 441, 291))
        self.formgroup.setFont(QFont("Times New Roman", 18))

        self.userlabel = QLabel("Username:", self.formgroup)
        self.userlabel.setGeometry(QRect(20, 40, 131, 51))

        self.passlabel = QLabel("Password:", self.formgroup)
        self.passlabel.setGeometry(QRect(20, 100, 131, 51))

        self.username = QLineEdit(self.formgroup)
        self.username.setGeometry(QRect(130, 49, 181, 31))
        self.username.setText("Sis Yemisi Adeshina")

        self.password = QLineEdit(self.formgroup)
        self.password.setGeometry(QRect(130, 110, 181, 31))
        self.password.setText("tr654")

        self.submit = QPushButton("Submit", self.formgroup)
        self.submit.setGeometry(QRect(134, 162, 171, 31))
        self.submit.setStyleSheet("background-color: rgb(0, 255, 0);")
        self.submit.clicked.connect(self.confirm)


        self.forgot = QPushButton("Forgot password", self.formgroup)
        self.forgot.setStyleSheet("background-color: orange; border-color: none")
        self.forgot.setGeometry(134, 200, 200, 31)
        self.forgot.clicked.connect(self.setup)
        self.password.setEchoMode(QLineEdit.Password)
        self.formgroup.show()

    def combo_changed(self):
        self.lp.setText(f"Phone Number({self.codes[str(self.residence.currentText())]})")

    def check(self):
        self.load = xl.load_workbook("children2.xlsx")
        self.pic = self.load["Sheet1"]
        self.info = eval(self.pic["A1"].value)
        try:
            for i in self.info.keys():
                if self.info[i] != self.username.text():
                    self.sety = True
                    pass
                else:
                    self.sety = False
                    QMessageBox.about(self, "Ërror", "The username you typed already exists")
                    break
            if self.sety:
                self.info[self.username.text()] = {"name":str(self.username.text()), "country":self.residence.currentText(), "phone":str(self.phoneno.text()), "email":str(self.email.text()), "password":str(self.password.text()), "date of birth":f"{self.calendar.selectedDate().month()}/{self.calendar.selectedDate().day()}/{self.calendar.selectedDate().year()}"}
                self.pic["A1"] = str(self.info)
                self.load.save("C:\\Users\\HP\\PycharmProjects\\untitled\\GenChamps\\children2.xlsx")
                self.welcomnewMember()
            else:
                pass
        except Exception as e:
            print(e)
            self.formgroup.show()
            QMessageBox.about(self, "Error", "There is an error in what you wrote.")

    def register(self):
        self.residence = QComboBox()
        for i in self.codes.keys():
            self.residence.addItem(i)
        self.residence.currentTextChanged.connect(self.combo_changed)

        self.vbox2 = QVBoxLayout()
        self.calendar = QCalendarWidget(self)
        self.calendar.setFont(QFont('Times', 6))
        self.calendar.setFixedHeight(200)
        date = QDate.currentDate()
        self.calendar.setMaximumDate(date)
        self.vbox2.addWidget(self.calendar)
        self.register_button.hide()
        self.login_button.hide()
        self.login_button2.hide()
        self.formgroup = QGroupBox("Register your child", self.widget)
        self.formgroup.setFont(QFont("Times New Roman", 15))
        self.formgroup.setGeometry(100, 100, 900, 700)
        self.box = QDialogButtonBox()
        self.box.setStandardButtons(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.box.accepted.connect(self.check)
        self.layout = QFormLayout()
        self.layout.setVerticalSpacing(30)
        self.phoneno = QLineEdit()
        self.username = QLineEdit()
        self.email = QLineEdit()
        self.password = QLineEdit()
        self.userpass = QLineEdit()
        self.submiti = QPushButton("Submit", self)
        self.submiti.setStyleSheet("background-color: green")
        self.submiti.clicked.connect(self.check)
        self.userpass.setEchoMode(QLineEdit.Password)
        self.password.setEchoMode(QLineEdit.Password)
        self.phoneno.setValidator(QIntValidator())
        my_regex = QRegExp("^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$")
        my_validator = QRegExpValidator(my_regex, self.email)
        self.email.setValidator(my_validator)
        my_regex2 = QRegExp("^[0-9]+")
        my_validator2 = QRegExpValidator(my_regex2, self.phoneno)
        self.phoneno.setValidator(my_validator2)
        self.lp = QLabel(f"Phone Number({self.codes['Afghanistan ']}): ")
        self.layout.addRow("User Name:", self.username)
        self.layout.addRow("Your Password:", self.userpass)
        self.layout.addRow("Parent's Email:", self.email)
        self.layout.addRow("Parent's Password:", self.password)
        self.layout.addRow("Date of Birth:", self.vbox2)
        self.layout.addRow("Country of residence", self.residence)
        self.layout.addRow(self.lp, self.phoneno)
        self.layout.addRow("", self.submiti)
        self.formgroup.setLayout(self.layout)
        self.formgroup.show()
        #self.scroll.setWidgetResizable(False)


    def calculate_action(self):
        # getting birth date day
        birth = self.calendar.selectedDate()

        # getting year and month day of birth day
        birth_year = birth.year()
        birth_month = birth.month()
        birth_day = birth.day()

        # getting today date
        current = QDate.currentDate()
        # getting year and month day of current day
        current_year = current.year()
        current_month = current.month()
        current_day = current.day()

        # coverting dates into date object
        birth_date = datetime.date(birth_year, birth_month, birth_day)
        current_date = datetime.date(current_year, current_month, current_day)

        # getting difference in both the dates
        difference = current_date - birth_date

        # getting days from the difference
        difference = difference.days

        # getting years from the difference
        years = difference / 365.2422

        # getting round value of years
        years = round(years)

        # setting this value with the help of label
        return years

    def exit1(self):
        # self.lkk.hide()
        self.child.hide()
        self.teacher.hide()
    def sectors(self):
        self.workbook = xl.load_workbook("children.xlsx")
        self.dj = self.workbook["Sheet1"]
        self.tobestored = eval(self.dj["A1"].value)
        try:
            self.infor = self.tobestored[self.solveduser.text()]
            if self.passkey.text() != self.infor["password"]:
                QMessageBox.about(self, "Error", "There is an error. It is either the user or the password that you entered that was invalid")
            else:
                self.straightteach()
        except Exception as exc:
            print(exc)
            QMessageBox.about(self, "Error",
                              "There is an error. It is either the user or the password that you entered that was invalid")
    def sectors2(self):
        self.workbook = xl.load_workbook("children2.xlsx")
        self.dj = self.workbook["Sheet1"]
        self.tobestored = eval(self.dj["A1"].value)
        try:
            self.infor = self.tobestored[self.username.text()]
            if self.passkey.text() != self.infor["password"]:
                QMessageBox.about(self, "Error", "There is an error. It is either the user or the password that you entered that was invalid")
            else:
                self.welcomnewMember()
        except Exception as e:
            print(e)
            QMessageBox.about(self, "Error",
                              "There is an error. It is either the user or the password that you entered that was invalid")


    def straightteach(self):
        self.setStyleSheet("")
        self.laying.hide()
        self.submits.hide()
        self.exits.hide()
        self.classe1 = self.infor["classe"]
        self.fonted = QFont("Times", 35)

        self.qlabel = QLabel(f"Welcome back {self.infor['name']}", self)
        self.qlabel.setFont(self.fonted)
        self.qlabel.setGeometry(300, 200, 500, 300)

        if self.classe1 == "Preschool":
            self.viewpre()
        if self.classe1 == "Elementary":
            self.viewele()
        if self.classe1 == "Preeteens":
            self.viewteen()

    def presccur(self):
        self.setStyleSheet("")
        self.group.hide()
        self.qlabel.hide()
        self.labels = QLabel(self)
        self.labels.setFixedSize(586, 586 + 86)
        self.labels.setPixmap(QPixmap(f"July/0001.jpg"))
        self.exitbht1 = QPushButton("Exit", self)
        self.exitbht1.setStyleSheet("background-color: black; color: white;")
        self.exitbht1.clicked.connect(self.fgct2)
        self.vbox.addWidget(self.exitbht1)

    def prescele(self):
        self.setStyleSheet("")
        self.group.hide()
        self.qlabel.hide()
        for i in range(1, 31):
            if len(str(i)) == 1:
                i2 = "0" + str(i)
            else:
                i2 = str(i)
            globals()['self.labels%s' % i] = QLabel(self)
            globals()['self.labels%s' % i].setFixedSize(586, 586 + 86)
            globals()['self.labels%s' % i].setPixmap(QPixmap(f"July/00{i2}.jpg"))
            self.vbox.addWidget(globals()['self.labels%s' % i])
        self.exitbht1 = QPushButton("Exit", self)
        self.exitbht1.setStyleSheet("background-color: black; color: white;")
        self.exitbht1.clicked.connect(self.fgct2)
        self.vbox.addWidget(self.exitbht1)

    def fgct2(self):
        for i in range(1, 31):
            globals()['self.labels%s' % i].hide()
        self.exitbht1.hide()
        self.qlabel.show()
        self.group.show()

    def viewpre(self):
        self.group = QGroupBox("Here's what you can do", self.widget)
        self.group.setGeometry(400, 300, 700, 900)
        self.group.showNormal()
        self.grid = QGridLayout()
        self.group.setLayout(self.grid)
        self.btu = QPushButton("View Curriculum", self.group)
        self.btu.setStyleSheet("background-color: green; border: none;")
        self.btu.setFont(QFont("Times", 35))
        self.btu.clicked.connect(self.presccur)

        self.btu2 = QPushButton("Do Assignments\n  (Coming Soon)", self.group)
        self.btu2.setStyleSheet("background-color: red; border: none;")
        self.btu2.setFont(QFont("Times", 35))

        self.btu3 = QPushButton("Play Games", self)
        self.btu3.setStyleSheet("background-color: red; border: none;")
        self.btu3.setFont(QFont("Times", 35))
        self.btu3.clicked.connect(self.playgame)

        self.grid.addWidget(self.btu, 0,0)
        self.grid.addWidget(self.btu2, 0,1)
        self.grid.addWidget(self.btu3, 1,0)

    def playgame(self):
        self.group.hide()
        self.qlabel.hide()
        if self.next_choice == "*":
            self.addQuestion("Who died for our sins", 2, ["Samuel", "Jonah", "Jesus", "Hamaan", "Daniel"], "Hebrews 9:28", False, 'z')
        if self.next_choice == '-x':
            self.addQuestion("Where did Jonah go to when God told him to go to Ninevah", 1, ["Israel", "Tarshish", "Jerusalem", "America", "Egypt"], False, 'l')
        self.vbox.addWidget(self.nexter)
    def viewele(self):
        self.group = QGroupBox("Here's what you can do", self.widget)
        self.group.setStyleSheet("background-color: blue;")
        self.group.setGeometry(400, 300, 700, 900)
        self.group.showNormal()
        self.grid = QGridLayout()
        self.group.setLayout(self.grid)
        self.btu = QPushButton("View Curriculum", self)
        self.btu.setFont(QFont("Times", 35))
        self.btu2 = QPushButton("Do Assignments", self)
        self.btu2.setFont(QFont("Times", 35))
        self.btu3 = QPushButton("Play Games", self)
        self.btu3.setFont(QFont("Times", 35))
        self.grid.addWidget(self.btu, 0, 0)
        self.grid.addWidget(self.btu2, 0, 1)
        self.grid.addWidget(self.btu3, 1, 0)
    def viewteen(self):
        self.group = QGroupBox("Here's what you can do", self.widget)
        self.group.setGeometry(400, 300, 700, 900)
        self.group.showNormal()
        self.grid = QGridLayout()
        self.group.setLayout(self.grid)
        self.btu = QPushButton("View Curriculum", self)
        self.btu.setFont(QFont("Times", 35))
        self.btu2 = QPushButton("Do assignments", self)
        self.btu2.setFont(QFont("Times", 35))
        self.btu3 = QPushButton("Play Games", self)
        self.btu3.setFont(QFont("Times", 35))
        self.grid.addWidget(self.btu, 0, 0)
        self.grid.addWidget(self.btu2, 0, 1)
        self.grid.addWidget(self.btu3, 1, 0)



    def exitlogger(self):
        self.laying.hide()
        self.submits.hide()
        self.exits.hide()

        self.login_button.show()
        self.register_button.show()
        self.login_button2.hide()
    def login(self):
        self.login_button.hide()
        self.login_button2.hide()
        self.register_button.hide()
        self.hher = QHBoxLayout()
        self.widgets = QWidget()
        self.widgets.setLayout(self.hher)

        self.laying = QGroupBox("Login", self.widget)
        self.laying.setGeometry(400, 200, 600, 300)
        self.laying.setFont(QFont("Times", 30))
        self.laying.show()

        self.forms = QFormLayout()
        self.solveduser = QLineEdit()
        self.passkey = QLineEdit()

        self.submits = QPushButton("Submit", self)
        self.submits.setStyleSheet("background-color: green;")
        self.submits.clicked.connect(self.sectors)
        self.submits.setFixedSize(100, 50)

        self.exits = QPushButton("Exit", self)
        self.exits.setStyleSheet("background-color: black; color: white;")
        self.exits.clicked.connect(self.exitlogger)
        self.exits.setFixedSize(100, 50)

        self.passkey.setEchoMode(QLineEdit.Password)

        self.forms.addRow("User Name:", self.solveduser)
        self.forms.addRow("Password:", self.passkey)
        self.forms.addRow("", self.submits)
        self.forms.addRow("", self.exits)

        self.laying.setLayout(self.forms)

    def login2(self):
        self.login_button.hide()
        self.login_button2.hide()
        self.register_button.hide()
        self.hher = QHBoxLayout()
        self.widgets = QWidget()
        self.widgets.setLayout(self.hher)

        self.laying = QGroupBox("Login", self.widget)
        self.laying.setFixedSize(800, 150)
        self.laying.showNormal()
        self.laying.setFont(QFont("Times", 30))

        self.forms = QFormLayout()
        self.username = QLineEdit()
        self.passkey = QLineEdit()

        self.submits = QPushButton("Submit", self)
        self.submits.setStyleSheet("background-color: green;")
        self.submits.clicked.connect(self.sectors2)
        self.submits.setFixedSize(100, 50)
        self.submits.showNormal()

        self.exits = QPushButton("Exit", self)
        self.exits.setStyleSheet("background-color: black; color: white;")
        self.exits.clicked.connect(self.exitlogger)
        self.exits.setFixedSize(100, 50)

        self.passkey.setEchoMode(QLineEdit.Password)

        self.forms.addRow("User Name:", self.username)
        self.forms.addRow("Password:", self.passkey)

        self.laying.setLayout(self.forms)

    def startMenu(self):
        self.setStyleSheet("")
        self.exit1()
        self.register_button = QPushButton("Register", self.widget)
        self.register_button.setFont(QFont("Times New Roman", 24))
        self.register_button.setStyleSheet("background-color: rgb(255, 0, 4);")
        self.register_button.setGeometry(QRect(290, 180, 191, 61))
        self.register_button.clicked.connect(self.register)
        self.register_button.show()

        self.login_button = QPushButton("Login\n(CE Member)", self.widget)
        self.login_button.clicked.connect(self.login)
        self.login_button.setFont(QFont("Times New Roman", 20))
        self.login_button.setGeometry(QRect(290, 260, 191, 61))
        self.login_button.setStyleSheet("background-color: qradialgradient(spread:repeat, cx:0.5, cy:0.5, radius:0.077, fx:0.5, fy:0.5, stop:0 rgba(0, 169, 255, 147), stop:0.497326 rgba(0, 0, 0, 147), stop:1 rgba(0, 169, 255, 147))")
        self.login_button.show()

        self.login_button2 = QPushButton("Login\n(Outreach Member)", self.widget)
        self.login_button2.clicked.connect(self.login2)
        self.login_button2.setGeometry(QRect(290, 340, 191, 61))
        self.login_button2.setFont(QFont("Times New Roman", 16))
        self.login_button2.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(0, 0, 0, 255), stop:0.05 rgba(14, 8, 73, 255), stop:0.36 rgba(28, 17, 145, 255), stop:0.6 rgba(126, 14, 81, 255), stop:0.75 rgba(234, 11, 11, 255), stop:0.79 rgba(244, 70, 5, 255), stop:0.86 rgba(255, 136, 0, 255), stop:0.935 rgba(239, 236, 55, 255));")
        self.login_button2.show()
        # self.vbox.addWidget(self.register_button)
        # self.vbox.addWidget(self.login_button)
        # self.vbox.addWidget(self.login_button2)


app = QApplication(sys.argv)
w = MainWindow()
w.show()
app.exec()